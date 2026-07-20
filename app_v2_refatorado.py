import streamlit as st
import pandas as pd
import os
import tempfile
import dre_engine_v2_refatorado as dre_engine

# ============================================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================================
st.set_page_config(page_title="DRE AI Dashboard", layout="wide", page_icon="✨")

st.markdown("""
<style>
.stApp { background-color: #0f172a; color: #f8fafc; }
h1, h2, h3 { color: #38bdf8; }
.stButton>button { background-color: #0ea5e9; color: white; border: none; border-radius: 8px; font-weight: bold; width: 100%; height: 50px; }
.stButton>button:hover { background-color: #0284c7; }
.alert-box { padding: 15px; border-radius: 8px; margin: 10px 0; }
.alert-warning { background-color: #7c2d12; border-left: 4px solid #ea580c; }
.alert-success { background-color: #1b4332; border-left: 4px solid #2d6a4f; }
.alert-error { background-color: #5a1a1a; border-left: 4px solid #dc2626; }
</style>
""", unsafe_allow_html=True)

st.title("✨ DRE Gerencial - Painel de Controle Inteligente")
st.markdown("Gerencie, audite e gere o seu DRE utilizando processamento automático.")

# ============================================================================
# SIDEBAR: CONFIGURAÇÃO DE ARQUIVOS
# ============================================================================
st.sidebar.header("📁 Configuração de Arquivos")
st.sidebar.markdown("---")

# Opção 1: Usar arquivos locais (se existirem)
usar_arquivos_locais = False
pagar_path = None
notas_path = None
template_path = None

# Verificar se os arquivos existem no diretório atual
dir_atual = os.getcwd()
arquivo_pagar_local = os.path.join(dir_atual, "BD CONTAS A PAGAR.xlsx")
arquivo_notas_local = os.path.join(dir_atual, "BD NOTAS.xlsx")
arquivo_template_local = os.path.join(dir_atual, "DRE_Final_Processado.xlsx")

if os.path.exists(arquivo_pagar_local) and os.path.exists(arquivo_notas_local) and os.path.exists(arquivo_template_local):
    usar_arquivos_locais = st.sidebar.checkbox(
        "✓ Usar arquivos locais (encontrados na pasta)",
        value=True,
        help="Se marcado, usará os arquivos BD CONTAS A PAGAR.xlsx, BD NOTAS.xlsx e DRE_Final_Processado.xlsx da pasta atual."
    )
    if usar_arquivos_locais:
        pagar_path = arquivo_pagar_local
        notas_path = arquivo_notas_local
        template_path = arquivo_template_local
        st.sidebar.success("✓ Arquivos locais carregados com sucesso!")

# Opção 2: Upload de arquivos
if not usar_arquivos_locais:
    st.sidebar.markdown("### Upload de Arquivos")
    
    uploaded_pagar = st.sidebar.file_uploader(
        "📊 BD CONTAS A PAGAR.xlsx",
        type=["xlsx"],
        help="Selecione o arquivo de Contas a Pagar"
    )
    uploaded_notas = st.sidebar.file_uploader(
        "📦 BD NOTAS.xlsx",
        type=["xlsx"],
        help="Selecione o arquivo de Notas Fiscais"
    )
    uploaded_template = st.sidebar.file_uploader(
        "📋 Template DRE (DRE_Final_Processado.xlsx ou similar)",
        type=["xlsx"],
        help="Selecione o template do DRE"
    )
    
    if uploaded_pagar and uploaded_notas and uploaded_template:
        # Salvar temporariamente os arquivos
        temp_dir = tempfile.mkdtemp()
        pagar_path = os.path.join(temp_dir, "BD CONTAS A PAGAR.xlsx")
        notas_path = os.path.join(temp_dir, "BD NOTAS.xlsx")
        template_path = os.path.join(temp_dir, "DRE_Final_Processado.xlsx")
        
        with open(pagar_path, "wb") as f:
            f.write(uploaded_pagar.getbuffer())
        with open(notas_path, "wb") as f:
            f.write(uploaded_notas.getbuffer())
        with open(template_path, "wb") as f:
            f.write(uploaded_template.getbuffer())
        
        st.sidebar.success("✓ Todos os arquivos carregados com sucesso!")
    else:
        st.sidebar.warning("⚠️ Por favor, carregue todos os três arquivos para continuar.")

# ============================================================================
# SEÇÃO 1: CALCULADORA E GERAÇÃO DO DRE
# ============================================================================
if pagar_path and notas_path and template_path:
    st.header("1. Calculadora e Geração do DRE", divider="blue")
    st.markdown("Esta ação lê as bases `BD NOTAS` e `BD CONTAS A PAGAR` e gera o documento oficial idêntico ao modelo.")

    col1, col2 = st.columns([1, 2])
    with col1:
        ano_selecionado = st.selectbox(
            "📅 Selecione o Ano:",
            options=[2024, 2025, 2026, 2027, 2028],
            index=2,
            key="ano_selecionado"
        )
        
        col_a, col_b = st.columns(2)
        with col_a:
            mes_inicio = st.slider("Mês Início do Comparativo:", 1, 12, 1, key="mes_inicio")
        with col_b:
            mes_fim = st.slider("Mês Fim do Comparativo:", 1, 12, 3, key="mes_fim")

        if st.button("🚀 PROCESSAR DRE + COMPARATIVO"):
            with st.spinner("Lendo planilhas pesadas e executando DRE Engine..."):
                try:
                    # Determinar o caminho de saída
                    output_path = os.path.join(os.path.dirname(template_path), "DRE_Final_Processado_NOVO.xlsx")
                    
                    # Processar DRE
                    resultado = dre_engine.processar_dre(
                        pagar_path=pagar_path,
                        notas_path=notas_path,
                        template_path=template_path,
                        output_path=output_path,
                        ano=ano_selecionado,
                        mes_inicio=mes_inicio,
                        mes_fim=mes_fim
                    )
                    
                    # Mostrar resultado
                    if resultado['sucesso']:
                        st.success("✓ DRE atualizado com sucesso!")
                        st.info(f"Salvo em: {output_path}")
                        
                        # Mostrar alertas de auditoria
                        if resultado['divergencias']:
                            st.markdown("### ⚠️ Alertas de Auditoria - Divergências Detectadas")
                            for mes, div in resultado['divergencias'].items():
                                meses_nomes = {1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"}
                                st.warning(f"**Mês {meses_nomes[mes]}:** Diferença de R$ {div['diferenca']:,.2f}\n- Contas a Pagar: R$ {div['contas_a_pagar']:,.2f}\n- DRE: R$ {div['dre']:,.2f}")
                        
                        if resultado['codigos_ausentes']:
                            st.markdown("### ⚠️ Alertas de Auditoria - Códigos de Centro de Custo Ausentes")
                            for cod in resultado['codigos_ausentes']:
                                st.warning(f"**Código ausente no DRE:** `{cod}`\nVerifique se este código existe na coluna C do DRE_Gerencial ou se há um erro de digitação na base de Contas a Pagar.")
                        
                        if not resultado['divergencias'] and not resultado['codigos_ausentes']:
                            st.markdown("### ✓ Conferência de Auditoria")
                            st.success("**Todos os valores conferem perfeitamente!** Diferença zero entre Contas a Pagar e DRE.")
                        
                        # Botão de Download
                        st.markdown("---")
                        st.markdown("### 📥 Download do DRE Processado")
                        with open(output_path, 'rb') as file:
                            st.download_button(
                                label="📊 Baixar DRE_Final_Processado_NOVO.xlsx",
                                data=file.read(),
                                file_name="DRE_Final_Processado_NOVO.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_dre_button"
                            )
                        st.markdown("Clique no botão acima para baixar a planilha processada.")
                    else:
                        st.error(f"Erro ao processar DRE: {resultado.get('mensagem', 'Erro desconhecido')}")
                        
                except Exception as e:
                    st.error(f"Erro ao processar: {str(e)}")
                    import traceback
                    st.error(traceback.format_exc())

    # ========================================================================
    # SEÇÃO 2: AUDITORIA RÁPIDA (DRILL-DOWN)
    # ========================================================================
    st.header("2. Auditoria Rápida (Drill-Down)", divider="blue")
    st.markdown("Use esta ferramenta para clicar e descobrir rapidamente **quais despesas e lançamentos compõem os valores** que aparecem no DRE.")

    # Caching Data Loaders
    @st.cache_data
    def load_data_for_audit():
        p = pd.read_excel(pagar_path)
        col_vencimento = p.columns[6]
        col_cc = p.columns[16]
        col_v = p.columns[7]
        p['Data_Calculo'] = pd.to_datetime(p[col_vencimento], errors='coerce')
        p[col_v] = pd.to_numeric(p[col_v], errors='coerce').fillna(0)
        p = p.dropna(subset=['Data_Calculo'])
        p['Mes_Num'] = p['Data_Calculo'].dt.month
        p['Ano_Num'] = p['Data_Calculo'].dt.year
        return p, col_cc, col_v

    @st.cache_data
    def load_notas_for_audit():
        n = pd.read_excel(notas_path)
        col_data_nf = 'dataemissao'
        col_total_nf = 'total'
        col_custo_total = 'custo_total'
        col_liquida_nf = 'VENDA LIQUIDA'
        
        n['Data_Calculo'] = pd.to_datetime(n[col_data_nf], errors='coerce')
        n = n.dropna(subset=['Data_Calculo'])
        n['Mes_Num'] = n['Data_Calculo'].dt.month
        n['Ano_Num'] = n['Data_Calculo'].dt.year
        return n, col_total_nf, col_custo_total, col_liquida_nf

    @st.cache_data
    def get_cc_mapping(template_path):
        import openpyxl
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb['DRE_Gerencial']
        mapping = {}
        for r in range(9, 250):
            lbl = ws.cell(row=r, column=2).value
            cc = ws.cell(row=r, column=3).value
            if cc and str(cc).strip() != '' and lbl and str(lbl).strip() != '':
                mapping[str(cc).strip()] = str(lbl).strip()
        return mapping

    # Create Audit Tabs
    tab_despesas, tab_faturamento = st.tabs([
        "🔍 Auditoria de Despesas (Contas a Pagar)", 
        "📦 Auditoria de Faturamento e CMV (Notas Fiscais)"
    ])

    # ========================================================================
    # TAB: AUDITORIA DE DESPESAS
    # ========================================================================
    with tab_despesas:
        with st.spinner("Carregando base de pagamentos..."):
            df_pagar, cc_col, v_col = load_data_for_audit()
            cc_mapping = get_cc_mapping(template_path)
        
        if df_pagar is not None:
            meses_nomes = {1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"}
            
            col_y, col_m, col_c = st.columns([1, 1, 2])
            with col_y:
                anos_disponiveis = sorted(df_pagar['Ano_Num'].dropna().unique().astype(int), reverse=True)
                # Default to 2026 if present
                default_y_idx = anos_disponiveis.index(2026) if 2026 in anos_disponiveis else 0
                ano_sel_desp = st.selectbox("Ano:", anos_disponiveis, index=default_y_idx, key="sb_ano_desp")
            with col_m:
                mes_sel_desp = st.selectbox("Mês:", range(1, 13), format_func=lambda x: meses_nomes[x], key="sb_mes_desp")
            with col_c:
                cc_uniques = [cc for cc in df_pagar[cc_col].dropna().unique() if str(cc).strip() in cc_mapping]
                cc_uniques = sorted(cc_uniques, key=lambda x: cc_mapping.get(str(x).strip(), '').upper())
                cc_selecionado = st.selectbox(
                    "Centro de Custo no DRE:", 
                    cc_uniques, 
                    format_func=lambda x: f"{x} - {cc_mapping.get(str(x).strip(), '')}",
                    key="sb_cc"
                )
                
            filtered = df_pagar[(df_pagar['Ano_Num'] == ano_sel_desp) & (df_pagar['Mes_Num'] == mes_sel_desp) & (df_pagar[cc_col] == cc_selecionado)]
            total = filtered[v_col].sum()
            
            st.metric(
                label=f"Total no DRE em {meses_nomes[mes_sel_desp]}/{ano_sel_desp} para: {cc_mapping.get(str(cc_selecionado).strip(), '')}", 
                value=f"R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            
            st.markdown("### 📋 Lançamentos que compõem este valor:")
            if not filtered.empty:
                display_df = filtered[['Vencimento', 'NomeFor', 'Descricao', 'Valor', 'Situacao', 'Descricao_tipo_pagamento']].copy()
                display_df.columns = ['Vencimento', 'Fornecedor', 'Histórico/Descrição', 'Valor (R$)', 'Situação', 'Forma de Pagamento']
                display_df['Vencimento'] = pd.to_datetime(display_df['Vencimento']).dt.strftime('%d/%m/%Y')
                
                st.dataframe(
                    display_df, 
                    use_container_width=True,
                    column_config={
                        "Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")
                    }
                )
            else:
                st.info(f"Nenhum lançamento encontrado para este Centro de Custo em {meses_nomes[mes_sel_desp]}/{ano_sel_desp}.")

    # ========================================================================
    # TAB: AUDITORIA DE FATURAMENTO / CMV
    # ========================================================================
    with tab_faturamento:
        with st.spinner("Carregando base de notas fiscais..."):
            df_notas, col_total_nf, col_custo_total, col_liquida_nf = load_notas_for_audit()
            
        if df_notas is not None:
            meses_nomes = {1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"}
            
            col_y_nf, col_m_nf, col_mode = st.columns([1, 1, 2])
            with col_y_nf:
                anos_disp_nf = sorted(df_notas['Ano_Num'].dropna().unique().astype(int), reverse=True)
                default_y_nf_idx = anos_disp_nf.index(2026) if 2026 in anos_disp_nf else 0
                ano_sel_nf = st.selectbox("Ano:", anos_disp_nf, index=default_y_nf_idx, key="sb_ano_nf")
            with col_m_nf:
                mes_sel_nf = st.selectbox("Mês:", range(1, 13), format_func=lambda x: meses_nomes[x], key="sb_mes_nf")
            with col_mode:
                modo_selecionado = st.selectbox(
                    "Indicador para Drill-down:",
                    ["Faturamento Bruto (Total NF)", "Venda Líquida", "Custo de Mercadoria Vendida (CMV)"],
                    key="sb_modo_nf"
                )
                
            # Select correct column based on mode
            if modo_selecionado == "Faturamento Bruto (Total NF)":
                audit_col = col_total_nf
            elif modo_selecionado == "Venda Líquida":
                audit_col = col_liquida_nf
            else:
                audit_col = col_custo_total
                
            filtered_nf = df_notas[(df_notas['Ano_Num'] == ano_sel_nf) & (df_notas['Mes_Num'] == mes_sel_nf)]
            total_nf_val = filtered_nf[audit_col].sum()
            
            st.metric(
                label=f"Total de {modo_selecionado} em {meses_nomes[mes_sel_nf]}/{ano_sel_nf}:",
                value=f"R$ {total_nf_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            
            st.markdown("### 📋 Detalhamento dos Itens das Notas Fiscais:")
            if not filtered_nf.empty:
                display_nf_df = filtered_nf[['dataemissao', 'numeroNF', 'Nome_Cliente', 'Descricao', 'quantidade', 'unitario', audit_col]].copy()
                display_nf_df.columns = ['Emissão', 'Nº Nota', 'Cliente', 'Produto/Item', 'Qtd', 'Unitário (R$)', f'{modo_selecionado} (R$)']
                display_nf_df['Emissão'] = pd.to_datetime(display_nf_df['Emissão']).dt.strftime('%d/%m/%Y')
                
                st.dataframe(
                    display_nf_df,
                    use_container_width=True,
                    column_config={
                        "Unitário (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                        f'{modo_selecionado} (R$)': st.column_config.NumberColumn(format="R$ %.2f")
                    }
                )
            else:
                st.info(f"Nenhuma nota fiscal emitida em {meses_nomes[mes_sel_nf]}/{ano_sel_nf}.")

else:
    st.warning("⚠️ Por favor, configure os arquivos na barra lateral para continuar.")
