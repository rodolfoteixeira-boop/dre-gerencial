import pandas as pd
import openpyxl
from copy import copy
import os

def sanitizar_codigo_cc(codigo):
    """
    Sanitiza códigos de centro de custo para garantir que estejam no formato correto.
    Se o código for apenas o nível pai (ex: '08'), tenta encontrar um código correspondente
    no DRE ou retorna o código original.
    """
    if codigo is None:
        return None
    codigo_str = str(codigo).strip()
    return codigo_str

def conferir_divergencias(pagar_path, dre_path, ano=2026):
    """
    Realiza conferência de divergências entre Contas a Pagar e DRE.
    Retorna um dicionário com as divergências encontradas por mês.
    """
    # Carregar Contas a Pagar
    df_cp = pd.read_excel(pagar_path)
    col_cc = df_cp.columns[16]  # Coluna Q
    col_vencimento = df_cp.columns[6]  # Coluna G
    col_valor = df_cp.columns[7]  # Coluna H
    
    df_cp['Data_Calculo'] = pd.to_datetime(df_cp[col_vencimento], errors='coerce')
    df_cp = df_cp.dropna(subset=['Data_Calculo'])
    df_cp['Ano'] = df_cp['Data_Calculo'].dt.year
    df_cp['Mes'] = df_cp['Data_Calculo'].dt.month
    df_cp_ano = df_cp[df_cp['Ano'] == ano].copy()
    
    # Carregar DRE
    wb = openpyxl.load_workbook(dre_path, data_only=True)
    ws = wb['DRE_Gerencial']
    
    # Mapear códigos de CC no DRE (coluna C)
    dre_codes = {}
    for r in range(9, 250):
        cc_val = ws.cell(row=r, column=3).value
        if cc_val is not None and str(cc_val).strip() != '':
            cc_str = str(cc_val).strip()
            dre_codes[cc_str] = r
    
    month_col_idx = {1: 4, 2: 5, 3: 6, 4: 7, 5: 8, 6: 9, 7: 10, 8: 11, 9: 12, 10: 13, 11: 14, 12: 15}
    
    divergencias = {}
    codigos_ausentes = set()
    
    for m in range(1, 13):
        total_cp = df_cp_ano[df_cp_ano['Mes'] == m][col_valor].sum()
        
        # Soma dos valores de CC no DRE para o mês m
        col_idx = month_col_idx[m]
        soma_dre = 0
        for r in range(9, 250):
            cc_val = ws.cell(row=r, column=3).value
            if cc_val is not None and str(cc_val).strip() != '' and r != 11:
                soma_dre += ws.cell(row=r, column=col_idx).value or 0
        
        diff = abs(total_cp - soma_dre)
        if diff > 0.01:  # Ignorar diferenças menores que 1 centavo (arredondamento)
            divergencias[m] = {
                'contas_a_pagar': total_cp,
                'dre': soma_dre,
                'diferenca': total_cp - soma_dre
            }
    
    # Detectar códigos ausentes
    cc_cp_ano = set(df_cp_ano[col_cc].dropna().astype(str).str.strip().unique())
    codigos_ausentes = cc_cp_ano - set(dre_codes.keys())
    
    return {
        'divergencias': divergencias,
        'codigos_ausentes': codigos_ausentes,
        'dre_codes': dre_codes
    }

def processar_dre(pagar_path, notas_path, template_path, output_path, ano=2026, mes_inicio=1, mes_fim=3):
    """
    Processa o DRE com caminhos dinâmicos e sanitização de dados.
    
    Args:
        pagar_path: Caminho para BD CONTAS A PAGAR.xlsx
        notas_path: Caminho para BD NOTAS.xlsx
        template_path: Caminho para o template DRE (pode ser relativo ou absoluto)
        output_path: Caminho para salvar o DRE processado
        ano: Ano a processar (padrão: 2026)
        mes_inicio: Mês inicial do comparativo (padrão: 1)
        mes_fim: Mês final do comparativo (padrão: 3)
    
    Returns:
        Dicionário com informações sobre o processamento (divergências, alertas, etc.)
    """
    print("Carregando bases de dados...")
    
    # Validar se os arquivos existem
    if not os.path.exists(pagar_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {pagar_path}")
    if not os.path.exists(notas_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {notas_path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {template_path}")
    
    # 1. Carregar Contas a Pagar
    pagar = pd.read_excel(pagar_path)
    col_vencimento = pagar.columns[6]  # G
    col_valor = pagar.columns[7]  # H
    col_cc = pagar.columns[16]  # Q
    
    pagar['Data_Calculo'] = pd.to_datetime(pagar[col_vencimento], errors='coerce')
    pagar[col_valor] = pd.to_numeric(pagar[col_valor], errors='coerce').fillna(0)
    pagar_valid = pagar.dropna(subset=['Data_Calculo']).copy()
    pagar_valid['Mes'] = pagar_valid['Data_Calculo'].dt.month
    pagar_valid['Ano'] = pagar_valid['Data_Calculo'].dt.year
    pagar_valid = pagar_valid[pagar_valid['Ano'] == ano]
    
    # Agregar para o DRE
    sum_pagar = pagar_valid.groupby([col_cc, 'Mes'])[col_valor].sum().reset_index()
    
    # 2. Carregar Notas Fiscais
    notas = pd.read_excel(notas_path)
    col_data_nf = notas.columns[10]  # K
    col_total_nf = notas.columns[5]  # F
    col_custo_total = notas.columns[6]  # G
    col_liquida_nf = notas.columns[39]  # AN - VENDA LIQUIDA
    
    notas['Data_Calculo'] = pd.to_datetime(notas[col_data_nf], errors='coerce')
    notas_valid = notas.dropna(subset=['Data_Calculo']).copy()
    notas_valid['Mes'] = notas_valid['Data_Calculo'].dt.month
    notas_valid['Ano'] = notas_valid['Data_Calculo'].dt.year
    notas_valid = notas_valid[notas_valid['Ano'] == ano]
    
    sum_notas = notas_valid.groupby('Mes').agg({col_total_nf: 'sum', col_liquida_nf: 'sum', col_custo_total: 'sum'}).reset_index()
    
    print("Processando Arquivo Excel...")
    wb = openpyxl.load_workbook(template_path)
    ws = wb['DRE_Gerencial']
    
    month_cols = {1: 4, 2: 5, 3: 6, 4: 7, 5: 8, 6: 9, 7: 10, 8: 11, 9: 12, 10: 13, 11: 14, 12: 15}
    
    # Sanitizar labels na coluna B (remover '=' acidentais)
    for r in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=r, column=2)
        if isinstance(cell_b.value, str) and cell_b.value.startswith("="):
            cell_b.value = cell_b.value.lstrip("=")
    
    # Estrutura para capturar valores para o Resumo Comparativo
    month_data = {m: {} for m in range(1, 13)}

    # Limpar valores antigos
    from openpyxl.cell.cell import MergedCell
    for r in range(9, 250):
        cc_val = ws.cell(row=r, column=3).value
        if pd.notna(cc_val) and str(cc_val).strip() != '':
            for col_idx in month_cols.values():
                cell = ws.cell(row=r, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    # Injetar Faturamento e CMV
    for index, row in sum_notas.iterrows():
        m = int(row['Mes'])
        if m in month_cols:
            ws.cell(row=6, column=month_cols[m]).value = row[col_total_nf]
            ws.cell(row=7, column=month_cols[m]).value = row[col_liquida_nf]
            ws.cell(row=11, column=month_cols[m]).value = row[col_custo_total]
            
            month_data[m][6] = row[col_total_nf]
            month_data[m][7] = row[col_liquida_nf]
            month_data[m][11] = row[col_custo_total]

    # Injetar Despesas por C.Custo
    for r in range(9, 250):
        cc_val = ws.cell(row=r, column=3).value
        if pd.notna(cc_val) and str(cc_val).strip() != '':
            cc_str = str(cc_val).strip()
            subset = sum_pagar[sum_pagar[col_cc].astype(str).str.strip() == cc_str]
            for _, tr in subset.iterrows():
                m = int(tr['Mes'])
                if m in month_cols:
                    val = tr[col_valor]
                    ws.cell(row=r, column=month_cols[m]).value = val
                    month_data[m][r] = val
                    
    # Verificação Linhas 251/252/253
    for m in month_cols:
        col_idx = month_cols[m]
        tot = pagar_valid[pagar_valid['Mes'] == m][col_valor].sum()
        ws.cell(row=252, column=col_idx).value = tot
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        expense_refs = []
        for r_idx in range(9, 250):
            cc_val = ws.cell(row=r_idx, column=3).value
            if r_idx != 11 and pd.notna(cc_val) and str(cc_val).strip() != '':
                expense_refs.append(f"{col_letter}{r_idx}")
        new_formula = "=" + "+".join(expense_refs)
        ws.cell(row=253, column=col_idx).value = new_formula

    # ================= Preenchimento da Aba Resumo Comparável =================
    ws_comp = None
    for name in wb.sheetnames:
        if "Resumo Compar" in name:
            ws_comp = wb[name]
            break
    
    if ws_comp:
        print(f"Preenchendo Resumo Comparativo ({mes_inicio} a {mes_fim})...")
        
        def safe_write(ws_obj, r, c, val, num_format=None):
            cell = ws_obj.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = val
                if num_format:
                    cell.number_format = num_format

        def get_total_group(m_idx, start_r, end_r):
            total = 0
            for r in range(start_r, end_r + 1):
                total += month_data.get(m_idx, {}).get(r, 0)
            return total

        comparative_mapping = {
            6: lambda m: month_data.get(m, {}).get(6, 0),
            8: lambda m: month_data.get(m, {}).get(7, 0),
            7: lambda m: (month_data.get(m, {}).get(6, 0) - month_data.get(m, {}).get(7, 0)),
            9: lambda m: (month_data.get(m, {}).get(11, 0) + month_data.get(m, {}).get(12, 0)),
            10: lambda m: (month_data.get(m, {}).get(7, 0) - (month_data.get(m, {}).get(11, 0) + month_data.get(m, {}).get(12, 0))),
            11: lambda m: get_total_group(m, 14, 52),
            12: lambda m: get_total_group(m, 56, 92),
            13: lambda m: get_total_group(m, 96, 111),
            14: lambda m: get_total_group(m, 115, 127),
            15: lambda m: get_total_group(m, 131, 156),
            16: lambda m: get_total_group(m, 161, 164),
            18: lambda m: get_total_group(m, 171, 182),
            19: lambda m: get_total_group(m, 187, 213),
            20: lambda m: get_total_group(m, 218, 220),
            21: lambda m: get_total_group(m, 225, 228),
        }
        
        safe_write(ws_comp, 18, 2, "(-) Impostos pagos no mês")
        safe_write(ws_comp, 19, 2, "(-) Despesas Financeiras")
        safe_write(ws_comp, 20, 2, "(-) Remuneração dos Sócios")
        safe_write(ws_comp, 21, 2, "(-) Investimentos / CAPEX")
        safe_write(ws_comp, 22, 2, "LUCRO LÍQUIDO OPERACIONAL")

        meses_nomes = {1:"Jan", 2:"Fev", 3:"Mar", 4:"Abr", 5:"Mai", 6:"Jun", 7:"Jul", 8:"Ago", 9:"Set", 10:"Out", 11:"Nov", 12:"Dez"}
        header_text = f"{meses_nomes.get(mes_inicio, 'Jan')}-{meses_nomes.get(mes_fim, 'Mar')} {ano}"
        safe_write(ws_comp, 5, 3, header_text)
        
        for row_c in range(6, 25):
            safe_write(ws_comp, row_c, 3, None)
            safe_write(ws_comp, row_c, 4, None)

        final_totals = {}
        for row_c, calc_func in comparative_mapping.items():
            sum_period = 0
            for m in range(mes_inicio, mes_fim + 1):
                sum_period += calc_func(m)
            final_totals[row_c] = sum_period
            safe_write(ws_comp, row_c, 3, sum_period)

        op_expenses = sum(final_totals.get(r, 0) for r in [11, 12, 13, 14, 15, 16])
        ebitda = final_totals.get(10, 0) - op_expenses
        safe_write(ws_comp, 17, 3, ebitda)
        final_totals[17] = ebitda
        
        extra_expenses = sum(final_totals.get(r, 0) for r in [18, 19, 20, 21])
        lucro_liq = ebitda - extra_expenses
        safe_write(ws_comp, 22, 3, lucro_liq)
        final_totals[22] = lucro_liq

        rec_liq_periodo = final_totals.get(8, 0)
        if rec_liq_periodo and rec_liq_periodo != 0:
            for row_c in range(6, 23):
                val = ws_comp.cell(row=row_c, column=3).value
                if val is not None:
                    safe_write(ws_comp, row_c, 4, val / rec_liq_periodo, '0.00%')

    # ================= Criação da Aba Resumo Mensal =================
    print("Gerando Aba de Resumo Dinâmica...")
    if "Resumo DRE" in wb.sheetnames:
        del wb["Resumo DRE"]
    ws_resumo = wb.create_sheet("Resumo DRE")
    
    ws_resumo.column_dimensions['B'].width = 50
    for m in month_cols:
        c_letter = openpyxl.utils.get_column_letter(month_cols[m])
        ws_resumo.column_dimensions[c_letter].width = 15
    valid_starts = ("(-)", "SUBTOTAL", "TOTAL", "%", "RECEITA", "MARGEM", "EBITDA", "RESULTADO", "DEMONSTRAÇÃO")
    orig_sheet_name = ws.title
    current_res_row = 2
    for r in range(1, ws.max_row + 1):
        cell_orig = ws.cell(row=r, column=2)
        lbl = str(cell_orig.value).strip() if cell_orig.value else ""
        if lbl.upper().startswith(valid_starts):
            res_cell = ws_resumo.cell(row=current_res_row, column=2)
            res_cell.value = cell_orig.value
            if cell_orig.has_style:
                res_cell.font = copy(cell_orig.font)
                res_cell.fill = copy(cell_orig.fill)
                res_cell.border = copy(cell_orig.border)
                res_cell.alignment = copy(cell_orig.alignment)
            for m in month_cols:
                c_idx = month_cols[m]
                c_letter = openpyxl.utils.get_column_letter(c_idx)
                orig_val = ws.cell(row=r, column=c_idx)
                res_val = ws_resumo.cell(row=current_res_row, column=c_idx)
                res_val.value = f"='{orig_sheet_name}'!{c_letter}{r}"
                if orig_val.has_style:
                    res_val.font = copy(orig_val.font)
                    res_val.fill = copy(orig_val.fill)
                    res_val.border = copy(orig_val.border)
                    res_val.alignment = copy(orig_val.alignment)
                    res_val.number_format = copy(orig_val.number_format)
            current_res_row += 1
            
    # Configurações de impressão para A4 Paisagem
    ws_resumo.page_setup.orientation = ws_resumo.ORIENTATION_LANDSCAPE
    ws_resumo.page_setup.paperSize = ws_resumo.PAPERSIZE_A4
    ws_resumo.page_setup.fitToPage = True
    ws_resumo.page_setup.fitToWidth = 1
    ws_resumo.page_setup.fitToHeight = 1
    
    # Reduzir margens para aproveitar o máximo de espaço na folha única
    ws_resumo.page_margins.top = 0.25
    ws_resumo.page_margins.bottom = 0.25
    ws_resumo.page_margins.left = 0.25
    ws_resumo.page_margins.right = 0.25
    ws_resumo.page_margins.header = 0.0
    ws_resumo.page_margins.footer = 0.0
    
    wb.save(output_path)
    print("DRE Finalizado.")
    
    # Realizar conferência de divergências
    print("\nRealizando conferência de divergências...")
    conferencia = conferir_divergencias(pagar_path, output_path, ano)
    
    resultado = {
        'sucesso': True,
        'output_path': output_path,
        'divergencias': conferencia['divergencias'],
        'codigos_ausentes': list(conferencia['codigos_ausentes']),
        'mensagem': ''
    }
    
    if conferencia['divergencias']:
        resultado['mensagem'] = f"⚠️ AVISO: Detectadas {len(conferencia['divergencias'])} divergências na conferência!"
        print(resultado['mensagem'])
        for mes, div in conferencia['divergencias'].items():
            print(f"  Mês {mes}: Diferença de R$ {div['diferenca']:,.2f}")
    
    if conferencia['codigos_ausentes']:
        resultado['mensagem'] += f"\n⚠️ AVISO: Detectados {len(conferencia['codigos_ausentes'])} códigos de Centro de Custo ausentes no DRE!"
        print(resultado['mensagem'])
        for cod in conferencia['codigos_ausentes']:
            print(f"  Código ausente: {cod}")
    
    if not conferencia['divergencias'] and not conferencia['codigos_ausentes']:
        resultado['mensagem'] = "✓ Conferência OK: Todos os valores conferem perfeitamente!"
        print(resultado['mensagem'])
    
    return resultado
